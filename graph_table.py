import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter  # 列幅調整用
from PIL import Image as PILImage  # Pillow を追加

plt.rcParams["font.family"] = "Meiryo"

class StudyDataProcessor:
    """ 勉強時間データを読み込み、統計データを処理するクラス """

    def __init__(self, file_path, sheet_name):
        """ 初期化: ファイルの存在チェック・Excelデータの読み込み """
        self.file_path = file_path
        self.sheet_name = sheet_name

        # 📌 charts.xlsx がない場合は新規作成、既存のものは削除
        if not os.path.exists("charts.xlsx"):
            wb = Workbook()
            wb.save("charts.xlsx")
        else:
            os.remove("charts.xlsx")
            wb = Workbook()
            wb.save("charts.xlsx")

        self.df = self.load_data()
        self.add_dates()
        self.calculate_stats()

    def load_data(self):
        """ Excelのデータを取得し、空白セルまでのデータを整理 """
        df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, usecols="B")
        last_valid_index = df.dropna(how="all").index[-1]
        return df.loc[:last_valid_index].copy()

    def add_dates(self):
        """ 日付データを読み込み、DataFrameに追加 """
        dates = pd.read_excel(self.file_path, sheet_name=self.sheet_name, usecols="A", skiprows=1, nrows=len(self.df), header=None)
        dates.columns = ["日付"]
        dates["日付"] = pd.to_datetime(dates["日付"] - 2, origin="1899-12-30", unit="D")
        self.df.insert(0, "日付", dates["日付"])
        self.df.set_index("日付", inplace=True)

    def calculate_stats(self):
        """ 週・月・年単位の統計を計算 """
        self.df_week = self.df.resample("W").agg(["median", "mean", "max", "min"])
        self.df_month = self.df.resample("ME").agg(["median", "mean", "max", "min"])
        self.df_year = self.df.resample("YE").agg(["median", "mean", "max", "min"])
        
        """ 最新データ・直近7日間・30日間の統計を計算 """
        self.df_latest = self.df.tail(1)
        self.df_last7 = self.df.tail(7)
        self.df_last30 = self.df.tail(30)

        # 最新データの順位を計算
        latest_value = self.df_latest.iloc[0, 0]
        rank = (self.df["勉強時間"] > latest_value).sum() + 1
        total_records = len(self.df)
        self.latest_rank = f"{rank} 位 / {total_records} 位"

class GraphPlotter:
    """ 統計データを可視化し、PNG として保存するクラス """

    def __init__(self, study_processor):
        self.study_df = study_processor.df
        self.df_week = study_processor.df_week
        self.df_month = study_processor.df_month
        self.df_latest = study_processor.df_latest
        self.df_last7 = study_processor.df_last7
        self.df_last30 = study_processor.df_last30
        self.df_all = study_processor.df  # 全データ統計用
        self.latest_rank = study_processor.latest_rank

    def generate_all_image(self):
        """ 📊 最新の統計データと推移グラフを PNG に統合し、1/2 サイズで保存 """
    
        # 📊 グラフ作成
        fig, axes = plt.subplots(3, 1, figsize=(10, 16))
    
        self.plot_stats(axes[0], self.df_all, "勉強時間の推移（全データ）",  "-", is_all_data=True)
        self.plot_stats(axes[1], self.df_month, "月単位の勉強時間", "--")
        self.plot_stats(axes[2], self.df_week, "週単位の勉強時間", ":")
    
        fig.suptitle("勉強時間の推移と統計データ（全期間・週・月）", fontsize=16)
        plt.tight_layout()
        plt.savefig("study.png", dpi=300, bbox_inches="tight")
        plt.close(fig)
    
        print("✅ 統合された統計データを study.png に保存しました！")
    
        # 📌 PNG画像を1/2サイズに縮小
        original_img = PILImage.open("study.png")
        new_size = (original_img.width // 4, original_img.height // 4)  # 1/2サイズに縮小
        resized_img = original_img.resize(new_size)
        resized_img.save("study_resized.png")  # 縮小版を保存
    
        print("✅ PNG画像を 1/4 サイズに縮小しました！")
    
        # 📌 Excel に画像を貼り付ける処理を追加
        wb = load_workbook("charts.xlsx")
        ws_chart = wb.create_sheet("勉強グラフ")
    
        # 📌 縮小した PNG画像をロード
        img = Image("study_resized.png")
    
        # 📌 画像を A1 に配置
        ws_chart.add_image(img, "A1")
    
        # 📌 Excelファイルを保存
        wb.save("charts.xlsx")
    
        print("✅ PNG画像を A1:S23 の範囲内に収めました！")


    def plot_stats(self, ax, df, title, linestyle, is_all_data=False):
        """ 各期間の統計データを描画 """
        if is_all_data:
            ax.plot(df.index, df["勉強時間"], label="全データ", linestyle="-", color="blue")
        else:
            ax.plot(df.index, df["勉強時間"]["median"], label="中央値", linestyle=linestyle)
            ax.plot(df.index, df["勉強時間"]["mean"], label="平均", linestyle=linestyle)
            ax.plot(df.index, df["勉強時間"]["max"], label="最大値", linestyle=linestyle)
            ax.plot(df.index, df["勉強時間"]["min"], label="最小値", linestyle=linestyle)

        ax.set_title(title)
        ax.legend()
        ax.grid(True)
        
    def plot_summary_table(self):
        """ 📌 統計データを「表データ」シートに保存し、列幅を最適化 """
        data = [
            ["項目", "値"],  # ヘッダー
            ["全データの中央値", self.df_all["勉強時間"].median()],
            ["全データの平均", self.df_all["勉強時間"].mean()],
            ["最新データの順位", self.latest_rank],
            ["直近7日間の中央値", self.df_last7["勉強時間"].median()],
            ["直近7日間の平均", self.df_last7["勉強時間"].mean()],
            ["直近30日間の中央値", self.df_last30["勉強時間"].median()],
            ["直近30日間の平均", self.df_last30["勉強時間"].mean()],
        ]

        """ 📌 charts.xlsx に表データとグラフを別シートで保存 """
        wb = load_workbook("charts.xlsx")
        ws_data = wb.create_sheet("勉強の表データ")

        # 📌 データをシートに書き込む
        for r_idx, row in enumerate(data, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws_data.cell(row=r_idx, column=c_idx, value=value)
                
        for col in ws_data.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 1) * 2
            ws_data.column_dimensions[column].width = adjusted_width
            
        # 📌 Excelファイルを保存
        wb.save("charts.xlsx")

        print("✅ charts.xlsx に表データを保存しました！")


# 📌 クラスを使って処理を実行
study_processor = StudyDataProcessor("record.xlsx", "studytime")
graph_plotter = GraphPlotter(study_processor)
graph_plotter.generate_all_image()
graph_plotter.plot_summary_table()

import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter  # 列幅調整用
from PIL import Image as PILImage  # Pillow を追加
import matplotlib.dates as mdates

# 📌 日本語フォント設定
plt.rcParams["font.family"] = "Meiryo"

class TypingDataProcessor:
    """ タイピングデータを読み込み、統計計算とグラフ化を行う親クラス """

    def __init__(self, file_path, sheet_name, usecols, columns, image_name):
        """ 初期化: データの読み込み """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.columns = columns
        self.image_name = image_name
        self.df = self.load_data(usecols)
        self.add_dates()
        self.process_data()
        self.calculate_stats()
        self.calculate_recent_stats()


    def load_data(self, usecols):
        """ Excelデータを取得し、空白セルまでのデータを整理 """
        df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, usecols=usecols)
        last_valid_index = df.dropna(how="all").index[-1]
        return df.loc[:last_valid_index].copy()

    def add_dates(self):
            """ 日付データを読み込み、DataFrameに追加 """
            dates = pd.read_excel(self.file_path, sheet_name=self.sheet_name, usecols="A", skiprows=1, nrows=len(self.df), header=None)
            dates.columns = ["日付"]

            # 日付列の型をチェックし、数値（Excelのシリアル値）のみを変換する
            dates["日付"] = dates["日付"].apply(
                lambda x: pd.to_datetime(x - 2, origin="1899-12-30", unit="D")
                if isinstance(x, (float, int))
                else pd.to_datetime(x)
            )
            
            self.df.insert(0, "日付", dates["日付"])
            self.df.set_index("日付", inplace=True)

    def calculate_stats(self):
        self.df_week = self.df.resample("W").agg(["median", "mean", "max", "min"])
        self.df_month = self.df.resample("ME").agg(["median", "mean", "max", "min"])

    def calculate_recent_stats(self):
        """ 最新データ・直近14日間・60日間の統計を計算（各項目ごと） """
    
        # 📌 最新データ（最終2行）
        self.df_latest = self.df.tail(2)
    
        # 📌 直近14日間・60日間のデータ
        self.df_last14 = self.df.tail(14)
        self.df_last60 = self.df.tail(60)
    
        # 📌 直近14日間・60日間の統計を計算（各カラムごと）
        self.stats_last14 = self.df_last14[self.columns].agg(["median", "mean", "max", "min"])
        self.stats_last60 = self.df_last60[self.columns].agg(["median", "mean", "max", "min"])
    
        # 📌 最新データの順位（2つ取得）
        self.latest_rank = {}
        for col in self.columns:
            latest_value_1 = self.df_latest[col].iloc[0]  # 最新データ1
            latest_value_2 = self.df_latest[col].iloc[1]  # 最新データ2
            rank_1 = (self.df[col] > latest_value_1).sum() + 1
            rank_2 = (self.df[col] > latest_value_2).sum() + 1
            total_records = len(self.df)
            self.latest_rank[col] = [f"{rank_1} 位 / {total_records} 位", f"{rank_2} 位 / {total_records} 位"]
    
        print("📌 最新データ（2つ）:", self.df_latest)
        print("📌 直近14日間の統計:", self.stats_last14)
        print("📌 直近60日間の統計:", self.stats_last60)
        print("📌 最新順位:", self.latest_rank)

    def process_data(self):
        pass  # ✅ 継承先で個別処理を定義する

    def plot_graphs(self):
        """ 📊 グラフ作成: 縦3×横1の全データ推移 & 週・月比較（PNGとして保存） """
    
        # 📌 1. 週・月比較のグラフ（3 × 2 のレイアウト）
        fig, axes_week_month = plt.subplots(nrows=3, ncols=2, figsize=(12, 16), sharex=True)
    
        for i, col in enumerate(self.columns):
            # 📌 週単位の統計
            axes_week_month[i, 0].plot(self.df_week.index, self.df_week.loc[:, (col, "median")], label=f"{col}（中央値）", linestyle="--", color="blue")
            axes_week_month[i, 0].plot(self.df_week.index, self.df_week.loc[:, (col, "mean")], label=f"{col}（平均）", linestyle="-", color="green")
            axes_week_month[i, 0].plot(self.df_week.index, self.df_week.loc[:, (col, "max")], label=f"{col}（最大）", linestyle=":", color="red")
            axes_week_month[i, 0].plot(self.df_week.index, self.df_week.loc[:, (col, "min")], label=f"{col}（最小）", linestyle="-.", color="purple")
        
            axes_week_month[i, 0].set_title(f"週ごとの {col}")
            axes_week_month[i, 0].legend()
        
            # 📌 月単位の統計
            axes_week_month[i, 1].plot(self.df_month.index, self.df_month.loc[:, (col, "median")], label=f"{col}（中央値）", linestyle="--", color="blue")
            axes_week_month[i, 1].plot(self.df_month.index, self.df_month.loc[:, (col, "mean")], label=f"{col}（平均）", linestyle="-", color="green")
            axes_week_month[i, 1].plot(self.df_month.index, self.df_month.loc[:, (col, "max")], label=f"{col}（最大）", linestyle=":", color="red")
            axes_week_month[i, 1].plot(self.df_month.index, self.df_month.loc[:, (col, "min")], label=f"{col}（最小）", linestyle="-.", color="purple")
        
            axes_week_month[i, 1].set_title(f"月ごとの {col}")
            axes_week_month[i, 1].legend()
    
        # 📌 2. 全データ推移のグラフ（縦3×横1）
        fig_all, axes_all = plt.subplots(nrows=len(self.columns), ncols=1, figsize=(10, 12), sharex=True)
    
        for i, col in enumerate(self.columns):
            axes_all[i].plot(self.df.index, self.df[col], label=f"{col} 全データ", color="blue")
            axes_all[i].set_title(f"{col} の推移")
            axes_all[i].legend()
    
        # 📌 X軸のフォーマットを統一
        first_date = self.df.index.min()
        last_date = self.df.index.max()
        middle_date = self.df.index[int(len(self.df) / 2)]
        tick_dates = [first_date, middle_date, last_date]
    
        for ax in axes_week_month.flatten():
            ax.set_xticks(tick_dates)
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y/%m/%d"))
    
        for ax in axes_all:
            ax.set_xticks(tick_dates)
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y/%m/%d"))
    
        fig.suptitle(f"{self.sheet_name} タイピングデータ", fontsize=16)
        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    
        # 📌 3. 両方のグラフを1つの画像に統合して保存
        fig.savefig("temp_week_month.png")
        img_week_month = PILImage.open("temp_week_month.png")
    
        fig_all.savefig("temp_all.png", bbox_inches="tight")
        img_all = PILImage.open("temp_all.png")
    
        new_height = img_week_month.height + img_all.height
        new_width = max(img_week_month.width, img_all.width)
        new_image = PILImage.new("RGB", (new_width, new_height), "white")
    
        new_image.paste(img_week_month, (0, 0))
        new_image.paste(img_all, (0, img_week_month.height))
    
        new_image.save(self.image_name)
    
        os.remove("temp_week_month.png")
        os.remove("temp_all.png")
    
        plt.close(fig)
        plt.close(fig_all)
        print(f"{self.image_name}に成功！")
    
        # 📌 Excel に画像を貼り付ける処理を追加
        wb = load_workbook("charts.xlsx")
        ws_chart = wb.create_sheet(f"{self.image_name} グラフ")
    
        # 📌 縮小した PNG画像をロード
        img = Image(f"{self.image_name}")
    
        # 📌 画像を A1 に配置
        ws_chart.add_image(img, "A1")
    
        # 📌 Excelファイルを保存
        wb.save("charts.xlsx")
        
    def plot_summary_table(self):
        """ 📊 各項目の統計情報（最新順位（2つ）・直近14日間・60日間・全データ）を**1枚の表にまとめて**表示＆保存 """
    
        headers = ["項目"] + self.columns
        data = []
    
        # 📌 統計データを作成（各カラムごとに値を揃える）
        metrics = {"median": "中央値", "mean": "平均", "min": "最小値", "max": "最大値"}
        periods = {"self.df": "全データ", "self.df_last14": "直近14日間", "self.df_last60": "直近60日間"}
    
        for df_key, period_name in periods.items():
            df_period = eval(df_key)  # `self.df`, `self.df_last14`, `self.df_last60` を取得
            for metric, metric_jp in metrics.items():
                row = [f"{period_name}の{metric_jp}"] + [df_period[col].agg(metric) for col in self.columns]
                data.append(row)
    
            data.append([""] * len(headers))  # 空行の区切り（正しく列数を揃える）
    
        # 📌 最新データの順位（2つ）
        rank_row_1 = ["最新データ順位（1つ目）"] + [self.latest_rank[col][0] for col in self.columns]
        rank_row_2 = ["最新データ順位（2つ目）"] + [self.latest_rank[col][1] for col in self.columns]
        data.append(rank_row_1)
        data.append(rank_row_2)
    
        data.append([""] * len(headers))  # 空行を追加


        """ 📌 charts.xlsx に表データとグラフを別シートで保存 """
        wb = load_workbook("charts.xlsx")
        ws_data = wb.create_sheet(f"{self.sheet_name}の表")
        
        # 📌 データをシートに書き込む
        for r_idx, row in enumerate(data, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws_data.cell(row=r_idx, column=c_idx, value=value)

        # 📌 列幅の最適化
        for col in ws_data.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 1) * 2
            ws_data.column_dimensions[column].width = adjusted_width

        # 📌 Excelファイルを保存
        wb.save("charts.xlsx")


class MyTypingProcessor(TypingDataProcessor):
    """ MyTyping の処理 """

    def __init__(self, file_path):
        columns = ["平均速度", "スコア", "正打率"]
        super().__init__(file_path, "mytyping", usecols="B:D", columns=columns, image_name="mytyping.png")

    def process_data(self):
        """ 正打率の計算 """
        self.df["正打率"] = self.df["正打率"] * 100

class SushidaProcessor(TypingDataProcessor):
    """ 寿司打 の処理 """

    def __init__(self, file_path, sheet_name):
        columns = ["総正打数", "スコア", "正打率"]
        image_name = f"{sheet_name}.png"  # ✅ シート名に応じてファイル名を変更
        super().__init__(file_path, sheet_name, usecols="B:D", columns=columns, image_name=image_name)

    def process_data(self):
        """ 正打率とスコアの計算 """
        self.df["正打率"] = (self.df["正しく打ったキー"] / (self.df["正しく打ったキー"] + self.df["ミス"]) * 100).round(1)
        self.df["スコア"] = ((self.df["正しく打ったキー"] - self.df["ミス"]) / (90 if self.sheet_name == "sushida5000" else 120) * 1000).round(0).astype(int)
        self.df["総正打数"] = self.df["正しく打ったキー"]
        self.df = self.df[["総正打数", "正打率", "スコア"]]

# 🎯 クラスの使い方
file_path = "record.xlsx"
processor_mytyping = MyTypingProcessor(file_path)
processor_mytyping.plot_graphs()  # ✅ mytyping.png が作成される
processor_mytyping.plot_summary_table()

processor_sushida_5000 = SushidaProcessor(file_path, "sushida5000")
processor_sushida_5000.plot_graphs()  # ✅ sushida.png が作成される
processor_sushida_5000.plot_summary_table()

processor_sushida_10000 = SushidaProcessor(file_path, "sushida10000")
processor_sushida_10000.plot_graphs()  # ✅ sushida.png が作成される
processor_sushida_10000.plot_summary_table()