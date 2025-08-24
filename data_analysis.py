import os
os.environ["OMP_NUM_THREADS"] = "3"

import mkl
mkl.set_num_threads(3)  # 🔹 MKL のスレッドを明示的に 3 に設定


import multiprocessing
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter  # 列幅調整用
from PIL import Image as PILImage  # Pillow を追加
import matplotlib.dates as mdates
from sklearn.linear_model import LinearRegression
from sklearn.cluster import KMeans
from sklearn.tree import DecisionTreeRegressor



# 📌 日本語フォント設定
plt.rcParams["font.family"] = "Meiryo"



class TypingDataProcessor:
    """ タイピングデータを読み込み、統計計算とグラフ化を行う親クラス """

    def __init__(self, file_path, sheet_name, usecols, columns, image_name):
        """ 初期化: データの読み込み """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.columns = columns
        self.image_name = image_name
        self.df = self.load_data(usecols)
        self.wb = None
        self.add_dates()
        self.process_data()
        self.calculate_stats()
        self.calculate_recent_stats()
    
    def make_excel(self):
        # "data_analysis.xlsx"を新しく作成
        if not os.path.exists("data_analysis.xlsx"):
            self.wb = Workbook()
            self.wb.save("data_analysis.xlsx")
        else:
            os.remove("data_analysis.xlsx")
            self.wb = Workbook()
            self.wb.save("data_analysis.xlsx")

    # 列名の切り分け
    def get_feature_columns_for_regression(self):
        """ 回帰分析に使う特徴量のカラム名を返す """
        if "総正打数" in self.df.columns:
            return ["総正打数", "正打率"]
        elif "平均速度" in self.df.columns:
            return ["平均速度", "正打率"]
        else:
            return ["正打率"]  # 最低限

    def get_target_column(self):
        return "スコア"
    
    def load_data(self, usecols):
        """ Excelデータを取得し、空白セルまでのデータを整理 """
        df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, usecols=usecols)
        last_valid_index = df.dropna(how="all").index[-1]
        return df.loc[:last_valid_index].copy()

    def add_dates(self):
            """ 日付データを読み込み、DataFrameに追加 """
            dates = pd.read_excel(self.file_path, sheet_name=self.sheet_name, usecols="A", skiprows=1, nrows=len(self.df), header=None)
            dates.columns = ["日付"]

            # 日付列の型をチェックし、数値（Excelのシリアル値）のみを変換する
            dates["日付"] = dates["日付"].apply(
                lambda x: pd.to_datetime(x - 2, origin="1899-12-30", unit="D")
                if isinstance(x, (float, int))
                else pd.to_datetime(x)
            )
            
            self.df.insert(0, "日付", dates["日付"])
            self.df.set_index("日付", inplace=True)

    def calculate_stats(self):
        self.df_week = self.df.resample("W").agg(["median", "mean", "max", "min"])
        self.df_month = self.df.resample("ME").agg(["median", "mean", "max", "min"])

    def calculate_recent_stats(self):
        """ 最新データ・直近14日間・60日間の統計を計算（各項目ごと） """
    
        # 📌 最新データ（最終2行）
        self.df_latest = self.df.tail(2)
    
        # 📌 直近14日間・60日間のデータ
        self.df_last14 = self.df.tail(14)
        self.df_last60 = self.df.tail(60)
    
        # 📌 直近14日間・60日間の統計を計算（各カラムごと）
        self.stats_last14 = self.df_last14[self.columns].agg(["median", "mean", "max", "min"])
        self.stats_last60 = self.df_last60[self.columns].agg(["median", "mean", "max", "min"])
    
        # 📌 最新データの順位（2つ取得）
        self.latest_rank = {}
        for col in self.columns:
            latest_value_1 = self.df_latest[col].iloc[0]  # 最新データ1
            latest_value_2 = self.df_latest[col].iloc[1]  # 最新データ2
            rank_1 = (self.df[col] > latest_value_1).sum() + 1
            rank_2 = (self.df[col] > latest_value_2).sum() + 1
            total_records = len(self.df)
            self.latest_rank[col] = [f"{rank_1} 位 / {total_records} 位", f"{rank_2} 位 / {total_records} 位"]
    
        print("📌 最新データ（2つ）:", self.df_latest)
        print("📌 直近14日間の統計:", self.stats_last14)
        print("📌 直近60日間の統計:", self.stats_last60)
        print("📌 最新順位:", self.latest_rank)

    def process_data(self):
        pass  # ✅ 継承先で個別処理を定義する

    def plot_graphs(self):
        """ 📊 グラフ作成: 縦3×横1の全データ推移 & 週・月比較（PNGとして保存） """
    
        # 📌 1. 週・月比較のグラフ（3 × 2 のレイアウト）
        fig, axes_week_month = plt.subplots(nrows=3, ncols=2, figsize=(12, 16), sharex=True)
    
        for i, col in enumerate(self.columns):
            # 📌 週単位の統計
            axes_week_month[i, 0].plot(self.df_week.index, self.df_week.loc[:, (col, "median")], label=f"{col}（中央値）", linestyle="--", color="blue")
            axes_week_month[i, 0].plot(self.df_week.index, self.df_week.loc[:, (col, "mean")], label=f"{col}（平均）", linestyle="-", color="green")
            axes_week_month[i, 0].plot(self.df_week.index, self.df_week.loc[:, (col, "max")], label=f"{col}（最大）", linestyle=":", color="red")
            axes_week_month[i, 0].plot(self.df_week.index, self.df_week.loc[:, (col, "min")], label=f"{col}（最小）", linestyle="-.", color="purple")
        
            axes_week_month[i, 0].set_title(f"週ごとの {col}")
            axes_week_month[i, 0].legend()
        
            # 📌 月単位の統計
            axes_week_month[i, 1].plot(self.df_month.index, self.df_month.loc[:, (col, "median")], label=f"{col}（中央値）", linestyle="--", color="blue")
            axes_week_month[i, 1].plot(self.df_month.index, self.df_month.loc[:, (col, "mean")], label=f"{col}（平均）", linestyle="-", color="green")
            axes_week_month[i, 1].plot(self.df_month.index, self.df_month.loc[:, (col, "max")], label=f"{col}（最大）", linestyle=":", color="red")
            axes_week_month[i, 1].plot(self.df_month.index, self.df_month.loc[:, (col, "min")], label=f"{col}（最小）", linestyle="-.", color="purple")
        
            axes_week_month[i, 1].set_title(f"月ごとの {col}")
            axes_week_month[i, 1].legend()
    
        # 📌 2. 全データ推移のグラフ（縦3×横1）
        fig_all, axes_all = plt.subplots(nrows=len(self.columns), ncols=1, figsize=(10, 12), sharex=True)
    
        for i, col in enumerate(self.columns):
            axes_all[i].plot(self.df.index, self.df[col], label=f"{col} 全データ", color="blue")
            axes_all[i].set_title(f"{col} の推移")
            axes_all[i].legend()
    
        # 📌 X軸のフォーマットを統一
        first_date = self.df.index.min()
        last_date = self.df.index.max()
        middle_date = self.df.index[int(len(self.df) / 2)]
        tick_dates = [first_date, middle_date, last_date]
    
        for ax in axes_week_month.flatten():
            ax.set_xticks(tick_dates)
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y/%m/%d"))
    
        for ax in axes_all:
            ax.set_xticks(tick_dates)
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y/%m/%d"))
    
        fig.suptitle(f"{self.sheet_name} タイピングデータ", fontsize=16)
        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    
        # 📌 3. 両方のグラフを1つの画像に統合して保存
        fig.savefig("temp_week_month.png")
        img_week_month = PILImage.open("temp_week_month.png")
    
        fig_all.savefig("temp_all.png", bbox_inches="tight")
        img_all = PILImage.open("temp_all.png")
    
        new_height = img_week_month.height + img_all.height
        new_width = max(img_week_month.width, img_all.width)
        new_image = PILImage.new("RGB", (new_width, new_height), "white")
    
        new_image.paste(img_week_month, (0, 0))
        new_image.paste(img_all, (0, img_week_month.height))
    
        new_image.save(self.image_name)
    
        os.remove("temp_week_month.png")
        os.remove("temp_all.png")
    
        plt.close(fig)
        plt.close(fig_all)
        print(f"{self.image_name}に成功！")
    
        # 📌 Excel に画像を貼り付ける処理を追加
        self.wb = load_workbook("data_analysis.xlsx")
        ws_data_analysis = self.wb.create_sheet(f"{self.image_name} グラフ")
    
        # 📌 縮小した PNG画像をロード
        img = Image(f"{self.image_name}")
    
        # 📌 画像を A1 に配置
        ws_data_analysis.add_image(img, "A1")
    
        # 📌 Excelファイルを保存
        self.wb.save("data_analysis.xlsx")
        
    def plot_summary_table(self):
        """ 📊 各項目の統計情報（最新順位（2つ）・直近14日間・60日間・全データ）を**1枚の表にまとめて**表示＆保存 """
    
        headers = ["項目"] + self.columns
        data = []
    
        # 📌 統計データを作成（各カラムごとに値を揃える）
        metrics = {"median": "中央値", "mean": "平均", "min": "最小値", "max": "最大値"}
        periods = {"self.df": "全データ", "self.df_last14": "直近14日間", "self.df_last60": "直近60日間"}
    
        for df_key, period_name in periods.items():
            df_period = eval(df_key)  # `self.df`, `self.df_last14`, `self.df_last60` を取得
            for metric, metric_jp in metrics.items():
                row = [f"{period_name}の{metric_jp}"] + [df_period[col].agg(metric) for col in self.columns]
                data.append(row)
    
            data.append([""] * len(headers))  # 空行の区切り（正しく列数を揃える）
    
        # 📌 最新データの順位（2つ）
        rank_row_1 = ["最新データ順位（1つ目）"] + [self.latest_rank[col][0] for col in self.columns]
        rank_row_2 = ["最新データ順位（2つ目）"] + [self.latest_rank[col][1] for col in self.columns]
        data.append(rank_row_1)
        data.append(rank_row_2)
    
        data.append([""] * len(headers))  # 空行を追加


        """ 📌 data_analysis.xlsx に表データとグラフを別シートで保存 """
        self.wb = load_workbook("data_analysis.xlsx")
        ws_data = self.wb.create_sheet(f"{self.sheet_name}の表")
        
        # 📌 データをシートに書き込む
        for r_idx, row in enumerate(data, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws_data.cell(row=r_idx, column=c_idx, value=value)

        # 📌 列幅の最適化
        for col in ws_data.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 1) * 2
            ws_data.column_dimensions[column].width = adjusted_width

        # 📌 Excelファイルを保存
        self.wb.save("data_analysis.xlsx")

    def analysis_trend_related(self):
        # 📌 Excelファイルの読み込み
        self.wb = load_workbook("data_analysis.xlsx")
        sheet_name = f"{self.image_name} 相関と回帰"
        ws_data_analysis = self.wb.create_sheet(sheet_name)
    
        # 🔹 **相関係数の出力（表形式）**
        ws_data_analysis.append(["相関係数"])
        correlation_matrix = self.df[self.columns].corr()
    
        # **ヘッダー追加**
        ws_data_analysis.append([""] + list(correlation_matrix.columns))
        
        # **行ごとのデータ追加**
        for index, row in correlation_matrix.iterrows():
            ws_data_analysis.append([index] + list(row))
    
        # **1行空ける**
        ws_data_analysis.append([])
    
        # 🔹 **データのばらつき**
        ws_data_analysis.append(["標準偏差"])
        std_values = {
            self.columns[0]: np.std(self.df[self.columns[0]]),
            self.columns[1]: np.std(self.df[self.columns[1]]),
            self.columns[2]: np.std(self.df[self.columns[2]])
        }
    
        for key, value in std_values.items():
            ws_data_analysis.append([key, value])
    
        # **1行空ける**
        ws_data_analysis.append([])
    
        # 🔹 **回帰分析**
        ws_data_analysis.append(["回帰分析"])
        
        X = self.df[[self.columns[0], self.columns[2]]]
        y = self.df[self.columns[1]]
 
        model = LinearRegression()
        model.fit(X, y)
    
        ws_data_analysis.append([f"{self.columns[0]} → {self.columns[1]}", model.coef_[0]])
        ws_data_analysis.append([f"{self.columns[2]} → {self.columns[1]}", model.coef_[1]])
        ws_data_analysis.append([f"{self.columns[1]}の決定係数 (R²)", model.score(X, y)])
    
        # **1行空ける**
        ws_data_analysis.append([])
    
        # 🔹 **正打率と総正打数の関係分析**
        rate_X = self.df[[self.columns[2]]]  
        rate_y = self.df[self.columns[0]]  
        rate_model = LinearRegression()
        rate_model.fit(rate_X, rate_y)
    
        ws_data_analysis.append([f"{self.columns[2]} → {self.columns[1]}", rate_model.coef_[0]])
        ws_data_analysis.append([f"{self.columns[2]}の決定係数 (R²)", rate_model.score(rate_X, rate_y)])
    
        # **1行空ける**
        ws_data_analysis.append([])
    
        # 🔹 **スコア予測**
        ws_data_analysis.append(["スコア予測"])
        new_data = pd.DataFrame({self.columns[0]: [600], self.columns[2]: [98]})
        predicted_score = model.predict(new_data)[0]
        ws_data_analysis.append(["予測スコア", predicted_score])
    
        # 📊 **ヒストグラムを作成**
        fig, axes = plt.subplots(3, 1, figsize=(8, 12))
        axes[0].hist(self.df[self.columns[1]], bins=5, alpha=0.5, color="blue")
        axes[0].set_title(f"{self.columns[1]}の分布")
        axes[1].hist(self.df[self.columns[2]], bins=5, alpha=0.5, color="green")
        axes[1].set_title(f"{self.columns[2]}の分布")
        axes[2].hist(self.df[self.columns[0]], bins=5, alpha=0.5, color="red")
        axes[2].set_title(f"{self.columns[0]}の分布")
    
        plt.tight_layout()
        graph_path = "hist.png"
        plt.savefig(graph_path, dpi=300, bbox_inches="tight")
        plt.close()
    
        # 📂 **グラフをExcelシートに追加**
        img = PILImage.open(graph_path)
        new_size = (img.width // 4, img.height // 4)
        resized_img = img.resize(new_size)
        resized_img.save(f"{graph_path}_resized.png")
        img = Image(f"{graph_path}_resized.png")
        ws_data_analysis.add_image(img, "A23")
        # 📌 Excelファイルを保存
        self.wb.save("data_analysis.xlsx")
        print(f"Excelファイル {sheet_name} にデータ＆グラフを保存しました！")

    def analysis_structure_features(self):
        os.environ["OMP_NUM_THREADS"] = "3"
        print(os.environ.get("OMP_NUM_THREADS"))  # 🔹 設定が反映されているかチェック
        print(multiprocessing.cpu_count())
        
        #移動平均
        self.df["スコア_7日平均"] = self.df[self.columns[1]].rolling(window=7).mean()
        self.df["スコア_30日平均"] = self.df[self.columns[1]].rolling(window=30).mean()
        
        self.df.plot(y=["スコア", "スコア_7日平均", "スコア_30日平均"], figsize=(10, 5))
        plt.title("スコアの移動平均")
        plt.close()

        #クラスター分析
        X = self.df[self.columns]
        model = KMeans(n_clusters=3, n_init=10)
        self.df["クラスタ"] = model.fit_predict(X)

        plt.scatter(self.df[self.columns[0]], self.df[self.columns[1]], c=self.df["クラスタ"], cmap="viridis")
        plt.xlabel(self.columns[0])
        plt.ylabel(self.columns[1])
        plt.title(f"スコアと{self.columns[0]}のクラスタ分類")
        plt.close()
        print(self.df.groupby("クラスタ").mean())

        plt.scatter(self.df[self.columns[2]], self.df[self.columns[1]], c=self.df["クラスタ"], cmap="viridis")
        plt.xlabel(self.columns[2])
        plt.ylabel(self.columns[1])
        plt.title(f"{self.columns[1]}と{self.columns[2]}のクラスタ分類")
        plt.close()
        print(self.df.groupby("クラスタ").mean())

        plt.scatter(self.df[self.columns[2]], self.df[self.columns[0]], c=self.df["クラスタ"], cmap="viridis")
        plt.xlabel(self.columns[2])
        plt.ylabel(self.columns[0])
        plt.title(f"{self.columns[2]}と{self.columns[0]}のクラスタ分類")
        plt.close()
        print(self.df.groupby("クラスタ").mean())

        #スコア向上の要因分析（特徴量重要度評価）
        X = self.df[self.get_feature_columns_for_regression()]
        y = self.df["スコア"]
        
        model = DecisionTreeRegressor()
        model.fit(X, y)

        print("特徴量の影響度:", model.feature_importances_)


class MyTypingProcessor(TypingDataProcessor):
    """ MyTyping の処理 """

    def __init__(self, file_path, sheet_name):
        columns = ["平均速度", "スコア", "正打率"]
        super().__init__(file_path, sheet_name, usecols="B:D", columns=columns, image_name="mytyping.png")
    def process_data(self):
        """ 正打率の計算 """
        self.df["正打率"] = self.df["正打率"] * 100

class SushidaProcessor(TypingDataProcessor):
    """ 寿司打 の処理 """

    def __init__(self, file_path, sheet_name):
        columns = ["総正打数", "スコア", "正打率"]
        image_name = f"{sheet_name}.png"  # ✅ シート名に応じてファイル名を変更
        super().__init__(file_path, sheet_name, usecols="B:D", columns=columns, image_name=image_name)

    def process_data(self):
        """ 正打率とスコアの計算 """
        self.df["正打率"] = (self.df["正しく打ったキー"] / (self.df["正しく打ったキー"] + self.df["ミス"]) * 100).round(1)
        self.df["スコア"] = ((self.df["正しく打ったキー"] - self.df["ミス"]) / (90 if self.sheet_name == "sushida5000" else 120) * 1000).round(0).astype(int)
        self.df["総正打数"] = self.df["正しく打ったキー"]
        self.df = self.df[["総正打数", "正打率", "スコア"]]

# sushida5000のクラス実行
file_path = "record.xlsx"
processor_sushida_5000 = SushidaProcessor(file_path, "sushida5000")
processor_sushida_5000.make_excel()
processor_sushida_5000.plot_graphs()  
processor_sushida_5000.plot_summary_table()
processor_sushida_5000.analysis_trend_related()
processor_sushida_5000.analysis_structure_features()

# sushida10000のクラス実行
processor_sushida_10000 = SushidaProcessor(file_path, "sushida10000")
processor_sushida_10000.plot_graphs()  
processor_sushida_10000.plot_summary_table()
processor_sushida_10000.analysis_trend_related()
processor_sushida_10000.analysis_structure_features()

# mytypingのクラス実行
processor_mytyping = MyTypingProcessor(file_path, "mytyping")
processor_mytyping.plot_graphs()  
processor_mytyping.plot_summary_table()
processor_mytyping.analysis_trend_related()
processor_mytyping.analysis_structure_features()